import json
import requests
import openpyxl
import os

from openpyxl.styles import PatternFill


def format_excel(file):
    key = os.environ.get('FNS_API_KEY')
    book = openpyxl.load_workbook(file)
    ws = book.active
    # TODO: uncomment when new key arrives
    # ogrns = [cell.value for cell in ws['H']][1:]
    # urls = [f'https://api-fns.ru/api/search?q={ogrn}&key={key}' for ogrn in ogrns]
    # print(urls)
    # companies = []
    # for url in urls:
    #     print(f'request for {url}')
    #     company = requests.get(url).json()
    #     print(f'got company {company}')
    #     companies.append(company)
    # print(companies)
    # with open('response.json', 'w', encoding='utf-8') as response_w:
    #     json.dump(companies, response_w, indent=4)
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for i in range(2, ws.max_row + 1):
        if ws[f'J{i}'].value == 'ЛИКВИДИРОВАНА':
            ws[f'K{i}'].value = 'НЕАКТИВНО'
            ws[f'K{i}'].fill = red_fill
    book.save('new.xlsx')
    book.close()


def main():
    format_excel('выверка МИНЦИФРЫ.xlsx')


if __name__ == '__main__':
    main()
